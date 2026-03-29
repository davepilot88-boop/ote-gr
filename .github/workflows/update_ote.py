import json
import re
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

PAGE_URL = "https://www.ote-cr.cz/cs/kratkodobe-trhy/elektrina/denni-trh"
ROOT = Path("data")
ROOT.mkdir(exist_ok=True)

def parse_date(s: str) -> datetime:
    return datetime.strptime(s, "%d.%m.%Y")

def find_xlsx_url(html: str) -> str:
    matches = re.findall(r'href="([^"]+\.xlsx[^"]*)"', html, flags=re.I)
    if not matches:
        raise SystemExit("Nenašel jsem XLSX odkaz.")
    for m in matches:
        if "DT_15MIN" in m.upper():
            return urljoin(PAGE_URL, m)
    return urljoin(PAGE_URL, matches[0])

def extract_prices(xlsx_bytes):
    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True)

    quarter_rows = []

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            values = [str(v).strip() if v else "" for v in row]
            if not values:
                continue

            if re.match(r"^\d{2}:\d{2}-\d{2}:\d{2}$", values[0]):
                nums = []
                for v in row:
                    if isinstance(v, (int, float)):
                        nums.append(float(v))
                    elif isinstance(v, str) and re.fullmatch(r"-?\d+[.,]\d+", v.strip()):
                        nums.append(float(v.replace(",", ".")))
                if nums:
                    quarter_rows.append(nums)

    if len(quarter_rows) >= 96:
        prices = [quarter_rows[i][-1] for i in range(0, 96, 4)]
        if len(prices) == 24:
            return prices

    raise SystemExit(f"Nenašel jsem 24 cen, jen {len(quarter_rows)} řádků.")

# --- MAIN ---

html = requests.get(PAGE_URL).text

m = re.search(r"(\d{2}\.\d{2}\.\d{4})", html)
if not m:
    raise SystemExit("Datum nenalezeno")

page_date = parse_date(m.group(1))

xlsx_url = find_xlsx_url(html)
xlsx_data = requests.get(xlsx_url).content

prices = extract_prices(xlsx_data)

payload = {
    "date": page_date.strftime("%d.%m.%Y"),
    "source": f"OTE XLSX",
    "prices": prices
}

today = datetime.now().date()
tomorrow = today + timedelta(days=1)

if page_date.date() == today:
    (ROOT / "today.json").write_text(json.dumps(payload, indent=2))
else:
    (ROOT / "tomorrow.json").write_text(json.dumps(payload, indent=2))

print("HOTOVO")
